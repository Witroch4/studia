"""
Consolida composições da pasta COMP-GEOPE (aba COMPOSIÇÃO_DESONERADA).

Layout GEOPE (diferente do "Composições próprias"):
  - Descrição: célula B6  -> [5,1]
  - Unidade:   célula J6  -> [5,9]
  - Código:    célula A7  -> [6,0]  (preenchido, ex.: "COMP01")

Saída: composicoes_geope_consolidadas.xlsx
"""
import os
import glob
import warnings
import xlrd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

warnings.filterwarnings("ignore")

GEOPE_DIR = "/home/wital/studia/CAGECE DEMANDA PLANILHA ORC/COMP-GEOPE"
OUT = "/home/wital/studia/composicoes_geope_consolidadas.xlsx"


def cell_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def extrair_geope():
    rows = []
    files = sorted(glob.glob(os.path.join(GEOPE_DIR, "*.xls")))
    for fp in files:
        fname = os.path.basename(fp)
        try:
            wb = xlrd.open_workbook(fp)
            # aceita "COMPOSIÇÃO_DESONERADA" ou variações
            aba = None
            for sn in wb.sheet_names():
                if "DESONERADA" in sn.upper():
                    aba = sn
                    break
            if not aba:
                print(f"  [skip] {fname}: sem aba DESONERADA")
                continue
            sh = wb.sheet_by_name(aba)

            descricao = cell_str(sh.cell_value(5, 1)) if sh.nrows > 5 and sh.ncols > 1 else ""
            unidade = cell_str(sh.cell_value(5, 9)) if sh.nrows > 5 and sh.ncols > 9 else ""
            codigo = cell_str(sh.cell_value(6, 0)) if sh.nrows > 6 and sh.ncols > 0 else ""

            if not descricao:
                print(f"  [warn] {fname}: descrição vazia")
                continue
            rows.append({
                "codigo": codigo,
                "descricao": descricao,
                "unidade": unidade,
                "arquivo": fname,
                "aba": aba,
            })
        except Exception as e:
            print(f"  [erro] {fname}: {e}")
    print(f"[OK] GEOPE: {len(rows)} extraídas de {len(files)} arquivos")
    return rows


def escrever_consolidado(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Composições GEOPE"

    headers = ["#", "Código", "Descrição", "Unidade", "Arquivo de Origem", "Aba de Origem"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="06B6D4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # linha de grupo
    n_row = ws.max_row + 1
    ws.cell(row=n_row, column=1, value=f"=== COMP-GEOPE ({len(rows)} itens) ===")
    ws.merge_cells(start_row=n_row, start_column=1, end_row=n_row, end_column=6)
    sep_cell = ws.cell(row=n_row, column=1)
    sep_cell.font = Font(bold=True, color="FFFFFF")
    sep_cell.fill = PatternFill("solid", fgColor="8B5CF6")
    sep_cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(rows, start=1):
        ws.append([
            i,
            r["codigo"],
            r["descricao"],
            r["unidade"],
            r["arquivo"],
            r["aba"],
        ])

    widths = [6, 18, 80, 12, 70, 32]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    wb.save(OUT)
    print(f"\n[OK] Arquivo gerado: {OUT}")
    print(f"[OK] Total: {len(rows)}")


def main():
    print("=" * 70)
    print("CONSOLIDANDO COMPOSIÇÕES CAGECE - GEOPE")
    print("=" * 70)
    rows = extrair_geope()
    escrever_consolidado(rows)


if __name__ == "__main__":
    main()
