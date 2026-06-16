"""Monta a planilha .xlsx do cronograma (estilo da planilha modelo ALECE)."""
from __future__ import annotations

import io
from datetime import date
from typing import Any

from openpyxl import Workbook

_DIAS_PT = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]
_FASE_LABEL = {
    "1volta": "1ª volta – resolver questões",
    "folga": "Folga/buffer",
    "buffer": "Buffer – revisão, erradas e simulados",
    "prova": "PROVA",
}


def montar_workbook(payload: dict[str, Any]) -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "Painel"
    ws["A1"] = f"Cronograma — {payload['nome_caderno']}"
    linhas = [
        ("Métrica", "Valor"),
        ("Data inicial", payload["data_inicio"]),
        ("Data da prova", payload["data_prova"]),
        ("Total de questões", payload["total"]),
        ("Questões resolvidas", "=COUNTIF(Cronograma!J:J,\">0\")"),
    ]
    for i, (a, b) in enumerate(linhas, start=3):
        ws.cell(i, 1, a)
        ws.cell(i, 2, b)

    cr = wb.create_sheet("Cronograma")
    headers = ["Data", "Dia", "Fase", "Questões novas", "Meta acumulada",
               "Feitas no dia", "Acumulado real", "Saldo", "Observações"]
    cr.append(headers)
    for r, d in enumerate(payload["plano"], start=2):
        cr.cell(r, 1, d.data)
        cr.cell(r, 2, _DIAS_PT[d.weekday])
        cr.cell(r, 3, _FASE_LABEL.get(d.fase, d.fase))
        cr.cell(r, 4, d.questoes_novas)
        cr.cell(r, 5, d.meta_acumulada)
        cr.cell(r, 7, f"=SUM($F$2:F{r})")
        cr.cell(r, 8, f"=G{r}-E{r}")

    di = wb.create_sheet("Discursivas")
    di.append(["Data", "Tema", "Tipo", "Qtd", "Status", "Nota", "Observações"])
    for x in payload.get("discursivas", []):
        di.append([x["data"], x["tema"], x.get("tipo", ""), x.get("qtd", 1),
                   x.get("status", "Pendente"), x.get("nota"), x.get("observacoes")])

    si = wb.create_sheet("Simulados")
    si.append(["Data", "Tipo", "Objetivas planejadas", "Meta objetiva",
               "Resultado objetiva", "Discursiva planejada", "Resultado discursiva"])
    for s in payload.get("simulados", []):
        si.append([s["data"], s["tipo"], s.get("objetivas_planejadas", 0),
                   s.get("meta_objetiva", 0), s.get("resultado_objetiva"),
                   s.get("discursiva_planejada", 0), s.get("resultado_discursiva")])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
