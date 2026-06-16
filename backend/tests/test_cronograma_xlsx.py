import io
from datetime import date

import openpyxl

from cronograma_core import gerar_plano, gerar_simulados


def test_montar_workbook_tem_abas_e_dados():
    plano = gerar_plano(date(2026, 5, 25), date(2026, 8, 16), 876, [6], 21)
    from cronograma_xlsx import montar_workbook
    payload = {
        "nome_caderno": "ALECE Eng Civil 876",
        "total": 876,
        "data_inicio": date(2026, 5, 25),
        "data_prova": date(2026, 8, 16),
        "plano": plano,
        "discursivas": [{"data": date(2026, 5, 26), "tema": "Fiscalização",
                         "tipo": "Treino 20 linhas", "qtd": 1, "status": "Pendente"}],
        "simulados": gerar_simulados(date(2026, 5, 25), date(2026, 8, 16), 21),
    }
    wb_bytes = montar_workbook(payload)
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))
    assert {"Painel", "Cronograma", "Discursivas", "Simulados"} <= set(wb.sheetnames)
    crono = wb["Cronograma"]
    assert crono["A1"].value == "Data"
    assert crono.max_row == len(plano) + 1
